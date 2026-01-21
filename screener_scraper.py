import time
import os
import re
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ================= EXCEL FORMATTING =================

def format_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active

    max_col = ws.max_column
    max_row = ws.max_row

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    particular_fill = PatternFill("solid", fgColor="E7F3FF")
    ratio_fill = PatternFill("solid", fgColor="FFFF00")

    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Header
    for col in range(1, max_col + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border

    # Rows that should be percentages (Excel row numbers)
    percentage_rows = {3, 5, 7, 8, 10, 17}

    for row in range(2, max_row + 1):
        first_cell = ws.cell(row=row, column=1)
        first_cell.fill = particular_fill
        first_cell.font = bold_font

        for col in range(1, max_col + 1):
            c = ws.cell(row=row, column=col)
            c.border = border
            if col > 1:
                c.alignment = center

                if row in percentage_rows:
                    c.number_format = "0.00%"
                    if isinstance(c.value, (int, float)):
                        c.value = c.value / 100
                elif isinstance(c.value, (int, float)):
                    c.number_format = "0.00"

    # Merge Ratios row
    ratio_row = 19
    ws.merge_cells(start_row=ratio_row, start_column=1, end_row=ratio_row, end_column=max_col)
    rc = ws.cell(row=ratio_row, column=1)
    rc.value = "Ratios"
    rc.fill = ratio_fill
    rc.font = Font(bold=True, size=12)
    rc.alignment = center
    rc.border = border

    # Auto width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 3

    ws.freeze_panes = "B2"
    wb.save(filename)


# ================= SCRAPER =================

class ScreenerExtractor:

    def __init__(self, username, password):
        self.username = username
        self.password = password
        self.driver = None
        self.ratios_added = False

    def start(self):
        options = Options()
        options.add_argument("--start-maximized")
        self.driver = webdriver.Chrome(options=options)

    def login(self):
        self.driver.get("https://www.screener.in/login/")
        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.NAME, "username"))).send_keys(self.username)
        self.driver.find_element(By.NAME, "password").send_keys(self.password)
        self.driver.find_element(By.XPATH, "//button[@type='submit']").click()
        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.ID, "desktop-search")))

    def search_company(self, company):
        box = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="desktop-search"]/div/input')))
        box.clear()
        box.send_keys(company)
        time.sleep(1)
        box.send_keys(Keys.SPACE)
        time.sleep(1)
        box.send_keys(Keys.ENTER)
        WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.ID, "profit-loss")))

    def get_number(self, xpath):
        try:
            el = WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.XPATH, xpath)))
            txt = re.sub(r"[^\d.\-]", "", el.text.replace(",", ""))

            return  float(txt)
        except:
            return 0.0

    def get_sales_cagr_5y(self):
        try:
            txt = self.driver.find_element(By.XPATH,
                "//*[@id='profit-loss']//table[1]//tr[td[contains(text(),'5 Years')]]/td[2]").text
            return float(txt.replace("%", ""))
        except:
            return 0.0

    def get_profit_cagr_5y(self):
        try:
            txt = self.driver.find_element(By.XPATH,
                "//*[@id='profit-loss']//table[2]//tr[td[contains(text(),'5 Years')]]/td[2]").text
            return float(txt.replace("%", ""))
        except:
            return 0.0

    def extract_core_financials(self):
        d = {}

        d["Revenue"] = self.get_number("//section[contains(.,'Profit & Loss')]//tr[td[contains(.,'Sales')]]/td[last()]")
        d["PAT"] = self.get_number("//section[contains(.,'Profit & Loss')]//tr[td[contains(.,'Net Profit')]]/td[last()]")
        d["Debt"] = self.get_number("//section[contains(.,'Balance Sheet')]//tr[td[contains(.,'Borrowings')]]/td[last()]")
        d["EBITDA"] = self.get_number("//section[contains(.,'Profit & Loss')]//tr[td[contains(.,'Operating Profit')]]/td[last()]")

        d["Networth"] = (
            self.get_number("//tr[td[contains(.,'Reserves')]]/td[last()]") +
            self.get_number("//tr[td[contains(.,'Equity Capital')]]/td[last()]")
        )

        d["Price"] = self.get_ratio("Current Price")
        d["CAGR_PAT_5Y"] = self.get_profit_cagr_5y()
        d["CAGR_Revenue_5Y"] = self.get_sales_cagr_5y()

        return d

    def add_ratio(self, name):
        box = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//input[@placeholder='eg. Promoter holding']")))
        box.send_keys(Keys.CONTROL + "a", Keys.DELETE)
        box.send_keys(name)
        time.sleep(1)
        box.send_keys(Keys.SPACE)
        time.sleep(1)
        box.send_keys(Keys.ENTER)
        time.sleep(1)

    def get_ratio(self, label):
        try:
            txt = self.driver.find_element(By.XPATH,
                f"//li[.//span[contains(text(),'{label}')]]//span[@class='number']").text
            return float(re.sub(r"[^\d.\-]", "", txt))
        except:
            return 0.0

    def extract_ratios(self):
        if not self.ratios_added:
            for r in ["No. of Equity Shares", "Market Cap", "EPS", "EPS growth 3Years",
                      "Piotroski score", "EVEBITDA", "PEG Ratio"]:
                self.add_ratio(r)
            self.ratios_added = True

        return {
            "Shares": self.get_ratio("No. Eq. Shares"),
            "MarketCap": self.get_ratio("Market Cap"),
            "EPS": self.get_ratio("EPS"),
            "EPS_Growth_3Y": self.get_ratio("EPS growth 3Years"),
            "Piotroski": self.get_ratio("Piotroski score"),
            "EV_EBITDA": self.get_ratio("EVEBITDA"),
            "PEG": self.get_ratio("PEG")
        }

    def close(self):
        self.driver.quit()


# ================= EXCEL CREATION =================

def create_excel(all_data):
    particulars = [
        "Revenue (Cr)", "CAGR Revenue 5 Years", "EBITDA (Cr)", "EBITDA Margins",
        "PAT (Cr)", "PAT Margins", "CAGR PAT 5 Years",
        "Networth (Cr)", "RoNW (Annualised)", "Debt (Cr)", "Price as of 01.11.2025",
        "",
        "No of Shares (Cr)", "Mkt Cap", "EPS", "EPS Growth 3 Years", "Piotroski score",
        "",
        "PE", "P/S", "P/BV", "Debt to Equity", "EV/EBITDA", "PEG Ratio", "Debt to EBITDA"
    ]

    df = pd.DataFrame({"Particulars": particulars})

    for i, (company, d) in enumerate(all_data.items()):
        col = chr(66 + i)

        df[company] = [
            d["Revenue"],
            d["CAGR_Revenue_5Y"],
            d["EBITDA"],
            f"=({col}4/{col}2)",
            d["PAT"],
            f"=({col}6/{col}2)",
            d["CAGR_PAT_5Y"],
            d["Networth"],
            f"=({col}6/{col}9)",
            d["Debt"],
            d["Price"],
            "",
            d["Shares"],
            d["MarketCap"],
            d["EPS"],
            d["EPS_Growth_3Y"],
            d["Piotroski"],
            "",
            f"={col}15/{col}6",
            f"={col}15/{col}2",
            f"={col}15/{col}9",
            f"={col}11/{col}9",
            d["EV_EBITDA"],
            d["PEG"],
            f"={col}11/{col}4"
        ]

    filename = "All_Companies_Analysis.xlsx"
    df.to_excel(filename, index=False)
    format_excel(filename)
    print("✅ Excel created & formatted successfully")


# ================= MAIN =================

if __name__ == "__main__":

    SCREENER_USERNAME = os.getenv("SCREENER_USERNAME")
    SCREENER_PASSWORD = os.getenv("SCREENER_PASSWORD")


    COMPANIES = [
        "S P Apparels Ltd",
        "Pearl Global Industries Ltd",
        "Nitin Spinners Ltd",
        "Gokaldas Exports Ltd",
        "Siyaram Silk Mills Ltd"
    ]

    bot = ScreenerExtractor(USERNAME, PASSWORD)
    bot.start()
    bot.login()

    all_data = {}

    for company in COMPANIES:
        bot.search_company(company)
        data = bot.extract_core_financials()
        data.update(bot.extract_ratios())
        all_data[company] = data

    bot.close()
    create_excel(all_data)
